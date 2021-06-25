from Bases_base import *
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import time
#https://www.delftstack.com/pt/howto/python-pandas/how-to-check-if-nan-exisits-in-pandas-dataframe/#pandas-dataframe-isull-m%C3%A9todo
#https://datenworks.com/transformacao-de-dados-em-python-com-pandas/

#Baixar o livro
#https://por.small-business-tracker.com/introducing-pandas-dataframe-510130#menu-1

def GetInstanceData():

    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")
    
    # Para instance que os dados vão
    instance = Instance_Data()

    #informações de dados valor de linha -2 e valor de coluna -1
    instance.num_depots = dados.iloc[ 3, 2]
    instance.num_customers = dados.iloc[4, 2]
    instance.num_locations = instance.num_depots + instance.num_customers
    instance.num_vehicle_types = dados.iloc[11, 2]

    #linguagem Portugues Brasil
    if dados.iloc[13,2] == "Não":
        instance.open_vrp = True
    else:
        instance.open_vrp = False


    if dados.iloc[13, 2] == "Sim - pode fazer isso várias vezes":
        instance.multi_trip = True
    else:
        instance.multi_trip = False
        
    if dados.iloc[14, 2] == "Difícil":
        instance.soft_time_windows = False
    else:
        instance.soft_time_windows = True
        
    if dados.iloc[15, 2] == "Não":
        instance.backhauls = False
    else:
        instance.backhauls = True

    #https://www.knowledgehut.com/blog/programming/how-to-work-with-excel-using-python
    #https://living-sun.com/pt/python/710241-python-validate-if-a-sheet-exists-in-my-document-xls-python-xls.html
    Compat_vehicle = load_workbook(filename="Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm", read_only=True)   # open an Excel file and return a workbook
    
    if '3.1. Compatibilidade do veículo' in Compat_vehicle.sheetnames:
        instance.vehicle_location_incompatibility = True
    else:
        instance.vehicle_location_incompatibility = True

    
    return instance

 
def GetSolverOptions():
    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    #retorna solver_options
    solver_options = Solver_Option_Data()

    if dados.iloc[20, 2] == "sim":
        solver_options.warm_start = True
    else:
        solver_options.warm_start = True

    if dados.iloc[20, 2] == "sim":
        solver_options.status_updates = True
    else:
        solver_options.status_updates = True
    
    solver_options.CPU_time_limit = dados.iloc[22,2]

    #porque a taxa de remoção minima é 0.15 e a maxima 0.35
    solver_options.LNS_minimum_removal_rate = 0.15 #'0.1
    solver_options.LNS_maximum_removal_rate = 0.35 #'0.3

    fraction_of_optional_customers = 0.0 # tá tipo double
    i=0

    #precisa do valor Vertex_list.Num_depots

    if fraction_of_optional_customers < 0.33:
        solver_options.LNS_candidate_list_size = 1
    elif fraction_of_optional_customers < 0.66:
        solver_options.LNS_candidate_list_size = 2
    else:
        solver_options.LNS_candidate_list_size = 3


    return solver_options

def GetVehicleTypeData():

    i = 0
    j = 0
    k = 0

    d= 0
    base_vehicle_type = 0

    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    vehicle_type_list = Vehicle_Type_List_Data()
    vertex_list = GetVertexData()

    #linha -2 e coluna -1
    base_vehicle_type = dados.iloc[11, 2]

    vehicle_type_list.num_vehicle_type = base_vehicle_type * vertex_list.num_depots

    dados2 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="3. Veículos")
    
    #capacity, fixed_cost_per_trip, cost_per_unit_distance, duration_multiplier, number_available, work_start_time, 
    # distance_limit, driving_time_limit, working_time_limit, origin_base_id, return_base_id, type_id
    k = 0
    for i in range(0, vertex_list.num_depots):
        for j in range(0, base_vehicle_type):
            capacity = dados2.iloc[k, 2]
            fixed_cost_per_trip = dados2.iloc[k, 3]
            cost_per_unit_distance = dados2.iloc[k, 4]
            duration_multiplier = dados2.iloc[k, 5]
            distance_limit = dados2.iloc[k, 6]
            work_s_t = dados2.iloc[k, 7]
            #work_start_time = Time_Converter(dados2.iloc[k,7], interval = "mins")
            #driving_time_limit = Time_Converter(dados2.iloc[k,8], interval = "mins")
            #working_time_limit = Time_Converter(dados2.iloc[k,9], interval = "mins")
            #number_available = Time_Converter(dados2.iloc[k,12], interval = "mins")
            #origin_base_id = i
            #return_base_id = (dados2.iloc[k,11] +1)
            #type_id = j

    print(work_s_t)
            
    '''       
            #capacity, fixed_cost_per_trip, cost_per_unit_distance, duration_multiplier, number_available, work_start_time, distance_limit, driving_time_limit, working_time_limit, origin_base_id, return_base_id, type_id
            vehicle_type_list.create_Vehicle_Type_List_Data(capacity, fixed_cost_per_trip, cost_per_unit_distance, 
            duration_multiplier, number_available, work_start_time, distance_limit, driving_time_limit , working_time_limit, 
            origin_base_id, return_base_id, type_id)

            k = k +1

    instance = GetInstanceData()

    if instance.vehicle_location_incompatibility == True:
        dados3 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="3.1. Compatibilidade do veículo")
        vehicle_type_list.capacity = [[0]*vertex_list.num_locations for i in range(vehicle_type_list.num_vehicle_type)]
        for i in range(0, vertex_list.num_locations):
            k=0
            for d in range(0, vertex_list.num_depots):
                for j in range(0, base_vehicle_type):

                    if dados3.iloc[i, 2+j] == 'Compatível':
                        vehicle_type_list.capacity[i, k] = True
                    else:
                        vehicle_type_list.capacity[i, k] = False

                    k = k +1
    '''
    #print(vehicle_type_list)
    return vehicle_type_list





def Time_Converter(tempo, interval = "secs"):

    hora = tempo.hour
    minutos = tempo.minute
    segundos = tempo.second
    
    '''
    def yrs():
      return divmod(duration_in_s, yr_ct)[0]

    def days():
      return divmod(duration_in_s, day_ct)[0]
    
    def hrs():
      return divmod(duration_in_s, hour_ct)[0]

    '''
    def mins():
        resutado  = (hora*60) + minutos + (segundos /60)
        return resutado
    '''
    def secs(): 
      return duration_in_s
    '''
    return {
        'mins': int(mins())
        #'secs': int(secs())
    }[interval]







def GetVertexData():
    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    #para testes
    vertex_list = Vertex_List_Data()

    vertex_list.num_depots = dados.iloc[3, 2]

    vertex_list.num_customers = dados.iloc[4, 2]
    vertex_list.num_locations = vertex_list.num_depots + vertex_list.num_customers

    #cria vetor estático
    dados2 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="1. Locais")
    i = 0
    
    #vertex_list.num_locations= 23
    for i in range(1, vertex_list.num_locations):
        tempo_inicio = dados2.iloc[i,6]
        time_window_start=Time_Converter(tempo_inicio, 'mins')
        vertex_list.create_Dataset('','', '','', '', time_window_start, '')
        #vertex_list.vertices.time_window_start = [Time_Converter(tempo_inicio, 'mins')]
    print(vertex_list.vertices[0])
 

#GetVertexData()
GetVehicleTypeData()
GetInstanceData()