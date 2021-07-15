from Bases_base import *
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings


#para ignorar avisos do modulo openpyxl porque a planinha está abreviada
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
#https://www.delftstack.com/pt/howto/python-pandas/how-to-check-if-nan-exisits-in-pandas-dataframe/#pandas-dataframe-isull-m%C3%A9todo
#https://datenworks.com/transformacao-de-dados-em-python-com-pandas/

#Baixar o livro
#https://por.small-business-tracker.com/introducing-pandas-dataframe-510130#menu-1

def Time_Converter(tempo):

    hora = tempo.hour
    minutos = tempo.minute
    segundos = tempo.second
    
    resutado  = (hora*60) + minutos + (segundos /60)
    
    return resutado


def GetVertexData():
    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    #para saber de qual classe pertence
    vertex_list = Vertex_List_Data()

    vertex_list.num_depots = dados.iloc[3, 2]

    vertex_list.num_customers = dados.iloc[4, 2]
    vertex_list.num_locations = vertex_list.num_depots + vertex_list.num_customers
    

    #cria vetor estático
    dados2 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="1. Locais")
    i = 0
    
    #foi coletado todos os dados da garagem e dos locais
    #vertex_list.num_locations= 23
    for i in range(0, vertex_list.num_locations):
        tempo_inicio = dados2.iloc[i,5]
        tempo_final = dados2.iloc[i,6]
        time_window_start = Time_Converter(tempo_inicio)
        time_window_end = Time_Converter(tempo_final)
        
        if dados2.iloc[i,7] == "Pode ser visitado":
            obrigado = 1
        elif dados2.iloc[i,7] == "Não visite":
            obrigado = -1
        else:
            obrigado = 0

        tempo_servico = dados2.iloc[i,8]
        service_time = Time_Converter(tempo_servico)
        Valor_de_retirada = dados2.iloc[i,9]
        Quantidade_de_entrega = dados2.iloc[i,10]
        Lucro = dados2.iloc[i,11]

        #service_time, mandatory, profit, pickup_amount, delivery_amount, time_windows_start, time_windows_end
        vertex_list.get_Vertex_Data(service_time, obrigado, Lucro, Valor_de_retirada, Quantidade_de_entrega, time_window_start, time_window_end)

    #objeto principal
    return vertex_list
    #print(vertex_list.vertices[1].time_windows_start)


vertex_list = GetVertexData()

def GetArcData():
    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="2. Distâncias")
    
    #para ver onde ser foram preenchido os dados
    arc_list = Arc_Data()
    #vertex_list = GetVertexData()
    
    arc_list.distance = np.zeros((vertex_list.num_locations, vertex_list.num_locations), dtype ='int16')
    arc_list.duration = np.zeros((vertex_list.num_locations, vertex_list.num_locations), dtype ='int16')
    i = 0
    j = 0
    k = 0
    for i in range(0, vertex_list.num_locations):
        for j in range(0, vertex_list.num_locations):
            #verificar a necessidade arredondar para baixo round(dados.iloc[k,2] - epsilon)
            arc_list.distance[i][j] = dados.iloc[k,2]
            arc_list.duration[i][j] = Time_Converter(dados.iloc[k,3])

            k = k + 1

    #print(arc_list.duration[0][3])
    #print(arc_list.duration[0,3])
    
    #Para saber que é o objeto principal 
    return arc_list


arc_list = GetArcData()


def GetInstanceData():

    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")
    
    # Para instance que os dados vão
    instance = Instance_Data()

    #informações de dados valor de linha -2 e valor de coluna -1
    instance.num_depots = dados.iloc[ 3, 2]
    instance.num_customers = dados.iloc[4, 2]
    instance.num_locations = instance.num_depots + instance.num_customers
    instance.num_vehicle_types = dados.iloc[11, 2]

    #linguagem Portugues Brasil
    if dados.iloc[13,2] == "Não":
        instance.open_vrp = True
    else:
        instance.open_vrp = False


    if dados.iloc[13, 2] == "Sim - pode fazer isso várias vezes":
        instance.multi_trip = True
    else:
        instance.multi_trip = False
        
    if dados.iloc[14, 2] == "Difícil":
        instance.soft_time_windows = False
    else:
        instance.soft_time_windows = True
        
    if dados.iloc[15, 2] == "Não":
        instance.backhauls = False
    else:
        instance.backhauls = True

    #https://www.knowledgehut.com/blog/programming/how-to-work-with-excel-using-python
    #https://living-sun.com/pt/python/710241-python-validate-if-a-sheet-exists-in-my-document-xls-python-xls.html
    Compat_vehicle = load_workbook(filename="Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm", read_only=True)   # open an Excel file and return a workbook
    
    if '3.1. Compatibilidade do veículo' in Compat_vehicle.sheetnames:
        instance.vehicle_location_incompatibility = True
    else:
        instance.vehicle_location_incompatibility = False

    #print(instance.vehicle_location_incompatibility)
    #Para saber que é o objeto principal

    
    


    return instance



def GetVehicleTypeData():

    i = 0
    j = 0
    k = 0

    d= 0
    #base_vehicle_types = 0

    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    #para saber de onde vem os dados
    vehicle_type_list = Vehicle_Type_List_Data()
    #vertex_list = GetVertexData()

    #linha -2 e coluna -1
    base_vehicle_types = dados.iloc[11, 2]
    

    #número de veículos * número de depositos
    vehicle_type_list.num_vehicle_types = base_vehicle_types * vertex_list.num_depots
    
    #https://question-it.com/questions/2208822/ignorirovat-userwarning-ot-openpyxl-s-pomoschju-pandas  
    dados2 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="3. Veículos")

    #capacity, fixed_cost_per_trip, cost_per_unit_distance, duration_multiplier, number_available, work_start_time, 
    # distance_limit, driving_time_limit, working_time_limit, origin_base_id, return_base_id, type_id
    k = 0
    for i in range(0, vertex_list.num_depots):
        for j in range(0, base_vehicle_types):
            capacity = dados2.iloc[k, 2]
            fixed_cost_per_trip = dados2.iloc[k, 3]
            cost_per_unit_distance = dados2.iloc[k, 4]
            duration_multiplier = dados2.iloc[k, 5]
            distance_limit = dados2.iloc[k, 6]
            work_s_t = dados2.iloc[k, 7]
            driving_t_l = dados2.iloc[k,8]
            working_t_l = dados2.iloc[k,9]
            work_start_time = Time_Converter(work_s_t)
            driving_time_limit = Time_Converter(driving_t_l)
            working_time_limit = Time_Converter(working_t_l)
            number_available = dados2.iloc[k,12]
            origin_base_id = i
            return_base_id = dados2.iloc[k, 11] #-> se salva da forma extendida
            #return_base_id = i
            type_id = j            

            #print(return_base_id)

            #capacity, fixed_cost_per_trip, cost_per_unit_distance, duration_multiplier, number_available, work_start_time, distance_limit, driving_time_limit, working_time_limit, origin_base_id, return_base_id, type_id
            vehicle_type_list.create_Vehicle_Type_List_Data(capacity, fixed_cost_per_trip, cost_per_unit_distance, 
            duration_multiplier, number_available, work_start_time, distance_limit, driving_time_limit , working_time_limit, 
            origin_base_id, return_base_id, type_id)
            k = k +1

    #print(vehicle_type_list.vehicle_types)

    #para saber de onde vem os dados
    instance =  GetInstanceData()
    
    if instance.vehicle_location_incompatibility == True:
        dados3 = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="3.1. Compatibilidade do veículo")
        vehicle_type_list.compatible = np.zeros((vertex_list.num_locations, vehicle_type_list.num_vehicle_types), dtype ='int16')
        for i in range(0, vertex_list.num_locations):
            k=0
            for d in range(0, vertex_list.num_depots):
                for j in range(0, base_vehicle_types):
                    if dados3.iloc[i, 2+j] == 'Compatível':
                        vehicle_type_list.compatible[i][k] = True
                    else:
                        vehicle_type_list.compatible[i][k] = False
                        
                    k = k +1
        
    

    #print(vehicle_type_list)
    #print(instance.vehicle_location_incompatibility)
    #objeto principal
    return vehicle_type_list


vehicle_type_list = GetVehicleTypeData()


def DeterminePenalty():
    
    instance =  GetInstanceData()
    
    distance_total = 0 #0.0 tipo doube
    cost_total = 0  #0.0 tipo doube
    
    i = 0
    j = 0

    #aqui é DeterminePenalty##################################################################:
    
    distance_total = 0 
    for i in range(0, vertex_list.num_locations):
        for j in range(0, vertex_list.num_locations):

            distance_total = distance_total + int(arc_list.distance[i,j])

    
    cost_total = 0
    for i in range(0, vehicle_type_list.num_vehicle_types):
        for j in range(0, vehicle_type_list.vehicle_types[i].NumberAvailable):
            cost_total = cost_total + vehicle_type_list.vehicle_types[i].CostPerUnitDistance

    if cost_total < 1:
        cost_total = 1

    for i in range(0, vehicle_type_list.num_vehicle_types):
        for j in range(0, vehicle_type_list.vehicle_types[i].NumberAvailable):
            instance.penalty = instance.penalty + vehicle_type_list.vehicle_types[i].FixedCostPerTrip
    
    
    #https://www.automateexcel.com/functions/ceiling-formula-excel/
    #Application.WorksheetFunction.Ceiling(instance.penalty, 1) -> transformando para inteiro
    instance.penalty = int(instance.penalty)


    if instance.penalty < 1000:
        instance.penalty = 1000

    ########DeterminePenalty#################################################################################:

    return instance


instance = DeterminePenalty()


def GetSolverOptions():
    dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")

    #retorna solver_options
    solver_options = Solver_Option_Data()

    if dados.iloc[20, 2] == "sim":
        solver_options.warm_start = True
    else:
        solver_options.warm_start = False

    if dados.iloc[20, 2] == "sim":
        solver_options.status_updates = True
    else:
        solver_options.status_updates = False
    
    solver_options.CPU_time_limit = dados.iloc[22,2]

    #porque a taxa de remoção minima é 0.15 e a maxima 0.35
    solver_options.LNS_minimum_removal_rate = 0.15 #'0.1
    solver_options.LNS_maximum_removal_rate = 0.35 #'0.3 #0.2 -> o outro algoritmo

    fraction_of_optional_customers = 0 #0.0 # tá tipo double
    
    #para ver onde ser foram preenchido os dados
    #vertex = GetVertexData()

    for i in range(vertex_list.num_depots, vertex_list.num_locations):
        if vertex_list.vertices[i].mandatory == 0:
            fraction_of_optional_customers = fraction_of_optional_customers +1
    
    fraction_of_optional_customers = fraction_of_optional_customers / vertex_list.num_customers
    #taxa de remoção testar com 0.2
    if fraction_of_optional_customers < 0.33:
        solver_options.LNS_candidate_list_size = 1
    elif fraction_of_optional_customers < 0.66:
        solver_options.LNS_candidate_list_size = 2
    else:
        solver_options.LNS_candidate_list_size = 3

    #print(fraction_of_optional_customers)
    #objeto principal
    return solver_options


solver_options = GetSolverOptions()


#dados = pd.read_excel('Dados/VRP_Spreadsheet_Solver_correta_Modelo.xlsm',sheet_name="VRP Solver Console")
#print(arc_list.distance)
#print(vertex_list.vertices[0].TimeWindowsEnd)
#print(vehicle_type_list.num_vehicle_types)